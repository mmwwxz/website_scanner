import socket
import ssl
import logging
import concurrent.futures
from datetime import datetime
from urllib.parse import urlparse
from functools import lru_cache, wraps
import os

import aiohttp
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')

GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def safe_execution(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            logging.error(f"Error in {func.__name__}: {e}")
            return None

    return wrapper


@lru_cache(maxsize=100)
def clean_url(url):
    parsed = urlparse(url)
    return parsed.netloc or parsed.path


def check_ports_multithreaded(host):
    open_ports = []
    common_ports = [443, 8080, 3000, 5000, 8000, 8081, 9000, 21, 22, 25, 53,
                    110, 143, 3306, 5432, 27017, 6379, 4002]

    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
        futures = {executor.submit(check_single_port, host, port): port for port in common_ports}
        for future in concurrent.futures.as_completed(futures):
            port = futures[future]
            try:
                result = future.result()
                if result:
                    open_ports.append(port)
            except Exception:
                pass

    return open_ports


def check_single_port(host, port):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(1)
        result = sock.connect_ex((host, port))
        sock.close()
        return result == 0
    except socket.error:
        return False


@safe_execution
def check_url(url, host, cache={}):
    if url in cache:
        return cache[url]

    try:
        response = requests.get(f"{url}", timeout=5, allow_redirects=True)

        if response.status_code == 404:
            soup = BeautifulSoup(response.content, "html.parser")
            title = soup.title.string if soup.title else ""

            interesting_patterns = ["page not found"]

            for pattern in interesting_patterns:
                if pattern in title.lower():
                    return {
                        'type': 'URL Check',
                        'host': host,
                        'details': f"Error 404 at {url}",
                        'status': 'OPEN'
                    }


            return {'type': 'URL Check', 'host': host, 'details': f"Error 404 at {url}", 'status': 'ERROR'}

        if response.status_code in [301, 302]:
            return {'type': 'URL Check', 'host': host,
                    'details': f"Redirected to {response.headers.get('Location')}", 'status': 'OPEN'}

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            title = soup.title.string if soup.title else ""

            interesting_patterns = [
                'admin', 'войти | административный сайт django',
                'swagger ui', 'api', 'openapi', 'documentation'
            ]

            for pattern in interesting_patterns:
                if pattern in title.lower():
                    return {
                        'type': 'URL Check',
                        'host': host,
                        'details': f"Interesting page found at {url} -> {title}",
                        'status': 'OPEN'
                    }

            return {'type': 'URL Check', 'host': host,
                    'details': f"Page available but not of special interest at {url} -> {title}",
                    'status': 'WARNING'}

        elif response.status_code >= 400:
            return {'type': 'URL Check', 'host': host,
                    'details': f"Error accessing {url}, status code {response.status_code}",
                    'status': 'ERROR'}

    except requests.exceptions.RequestException as e:
        return {'type': 'URL Check', 'host': host,
                'details': f"Error checking {url}: {e}", 'status': 'ERROR'}


@safe_execution
def check_ssl_expiry(host, port=443):
    try:
        context = ssl.create_default_context()
        with socket.create_connection((host, port), timeout=10) as sock:
            with context.wrap_socket(sock, server_hostname=host) as ssock:
                cert = ssock.getpeercert()
                expiry_date = datetime.strptime(cert['notAfter'], "%b %d %H:%M:%S %Y %Z")
                days_remaining = (expiry_date - datetime.utcnow()).days
                return {
                    'type': 'SSL Check',
                    'host': host,
                    'details': f"SSL certificate expires on {expiry_date}, {days_remaining} days remaining.",
                    'status': 'WARNING' if days_remaining < 30 else 'OPEN'
                }
    except Exception as e:
        return {'type': 'SSL Check', 'host': host,
                'details': f"Error checking SSL certificate for {host}: {e}", 'status': 'ERROR'}


def check_api_docs(host, port):
    possible_paths = [
        "/admin", "/administrator", "/admin/login", "/admin/dashboard", "/backend",
        "/cms", "/panel", "/console", "/secure/admin", "/admin-panel",
        "/config", "/config.php", "/settings.php", "/install.php", "/config/settings",
        "/setup", "/installer",
        "/api/auth", "/api/login", "/api/logout", "/oauth/token", "/api/users", "/api/admin",
        "/keys", "/certificates", "/token", "/tokens", "/secrets", "/api/secrets",
        "/db", "/database", "/dump", "/logs", "/log", "/debug.log", "/access.log", "/error.log",
        "/user", "/users", "/user/profile", "/payment", "/checkout", "/cart", "/invoice",
        "/order", "/orders", "/transactions", "/api/docs", "/api/redoc", '/api/asas'
    ]

    results = []
    for path in possible_paths:
        url_with_port = f"https://{host}:{port}{path}"
        url_without_port = f"https://{host}{path}"

        result_with_port = check_url(url_with_port, host)
        result_without_port = check_url(url_without_port, host)

        if result_with_port:
            results.append(result_with_port)
        if result_without_port:
            results.append(result_without_port)

    return results


def save_to_excel(results, filename):
    os.makedirs('document', exist_ok=True)

    full_path = os.path.join('document', filename)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Scan Results")

    ws.append(["Type", "Host", "Details", "Status"])

    for result in results:
        row = [result['type'], result['host'], result['details'], result['status']]
        ws.append(row)

    wb.save(full_path)
    logging.info(f"Results saved to {full_path}")
    return full_path


def scanner(host, output_filename=None):
    host = clean_url(host)
    results = []

    with concurrent.futures.ThreadPoolExecutor() as executor:
        port_future = executor.submit(check_ports_multithreaded, host)
        ssl_future = executor.submit(check_ssl_expiry, host)

        open_ports = port_future.result()
        ssl_result = ssl_future.result()

    if open_ports:
        for port in open_ports:
            port_results = check_api_docs(host, port)
            results.extend(port_results)
    else:
        results.append({
            'type': 'Port Check',
            'host': host,
            'details': 'No open ports found',
            'status': 'ERROR'
        })

    if ssl_result:
        results.append(ssl_result)

    if output_filename is None:
        output_filename = f"{host}_scan_results.xlsx"

    saved_path = save_to_excel(results, output_filename)

    return results, saved_path
